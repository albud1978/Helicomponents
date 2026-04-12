#!/usr/bin/env python3
"""
Нормализация новой папки датасета ``data_input/source_data/v_YYYY-MM-DD/`` до запуска extract.

- ``Program_AC.xlsx``: при необходимости переименовать заголовок ``direction`` → ``directorate``
  (первый лист, первая строка — как у pandas по умолчанию).
- ``Status_Components.xlsx``: колонка ``lease_restricted`` — правило по ``owner`` (str strip):
  ``'Y'`` для владельцев из ``LEASE_RESTRICTED_OWNERS``, иначе ``''`` (в Excel пустая ячейка;
  dual_loader мапит не-Y в 0).

  Список владельцев согласован с эмпирическими счётчиками по уже принятым датасетам
  (декабрь 2025 — март 2026); при смене бизнес-правил обновляйте frozenset и документацию.

Запуск extract после нормализации: ``python3 code/extract/extract_master.py`` (см. docs/architecture/extract.md).
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

# Согласовано с эмпирией по существующим датасетам (Dec 2025 – Mar 2026)
LEASE_RESTRICTED_OWNERS = frozenset({"ГТЛК", "ВТК-АВИА", "СБЕР ЛИЗИНГ"})

PROGRAM_AC_NAME = "Program_AC.xlsx"
STATUS_COMPONENTS_NAME = "Status_Components.xlsx"

# Предупреждение для очень больших xlsx (pandas полная перезапись)
HUGE_BYTES = 50 * 1024 * 1024
WARN_ROWCOUNT = 250_000


def _repo_root() -> Path:
    return Path(__file__).resolve().parent.parent.parent


def resolve_dataset_path(raw: str) -> Path:
    p = Path(raw).expanduser()
    if not p.is_absolute():
        p = _repo_root() / p
    return p.resolve()


def _fail(msg: str) -> None:
    print(f"ERROR: {msg}", file=sys.stderr)
    sys.exit(1)


def normalize_program_ac_header(
    path: Path, *, dry_run: bool
) -> None:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=False, data_only=False)
    ws = wb.worksheets[0]
    row1 = list(ws.iter_rows(min_row=1, max_row=1))[0]
    headers: list[tuple[object, str]] = []
    for cell in row1:
        v = cell.value
        s = str(v).strip() if v is not None else ""
        headers.append((cell, s))

    has_directorate = any(h[1].casefold() == "directorate" for h in headers)
    if has_directorate:
        print(f"  [Program_AC] OK: колонка «directorate» уже есть — шаг пропущен (idempotent).")
        wb.close()
        return

    to_rename = [h for h in headers if h[1].casefold() == "direction"]
    if not to_rename:
        _fail(
            f"{path}: нет колонок «directorate» и «direction» в первой строке первого листа — "
            "проверьте исходный Excel."
        )

    for cell, _ in to_rename:
        cell.value = "directorate"
    print(
        f"  [Program_AC] Переименование заголовка: direction → directorate "
        f"({len(to_rename)} ячеек)."
    )
    if dry_run:
        print("  [Program_AC] --dry-run: файл не записан.")
        wb.close()
        return
    wb.save(path)
    wb.close()
    print(f"  [Program_AC] Записано: {path}")


def _lease_series_from_owner(series) -> object:
    import numpy as np
    import pandas as pd

    owners = series.map(lambda x: "" if pd.isna(x) else str(x).strip())
    return np.where(owners.isin(LEASE_RESTRICTED_OWNERS), "Y", "")


def normalize_status_lease_restricted(
    path: Path, *, dry_run: bool, sync_lease: bool
) -> None:
    import pandas as pd

    size = path.stat().st_size
    if size >= HUGE_BYTES:
        print(
            f"  WARNING: {path.name} очень большой ({size / (1024 * 1024):.1f} MiB) — "
            "полная перезапись через pandas может занять заметное время.",
            file=sys.stderr,
        )

    header = pd.read_excel(path, engine="openpyxl", nrows=0)
    has_lease = "lease_restricted" in header.columns
    if has_lease and not sync_lease:
        print(
            "  [Status_Components] Колонка «lease_restricted» уже есть — шаг пропущен. "
            "Передайте --sync-lease для полного пересчёта по owner."
        )
        return

    if "owner" not in header.columns:
        _fail(f"{path}: для шага lease нужна колонка «owner».")

    df = pd.read_excel(path, engine="openpyxl")
    nrows = len(df)
    if nrows >= WARN_ROWCOUNT:
        print(
            f"  WARNING: {nrows:,} строк — убедитесь, что памяти достаточно.",
            file=sys.stderr,
        )

    new_lease = _lease_series_from_owner(df["owner"])
    if has_lease and sync_lease:
        print("  [Status_Components] --sync-lease: пересчёт всей колонки lease_restricted по owner.")
    else:
        print("  [Status_Components] Добавление колонки lease_restricted по правилу owner.")

    y_count = int((new_lease == "Y").sum())
    print(f"  [Status_Components] Строк с lease_restricted='Y': {y_count:,} (всего строк: {nrows:,}).")

    if dry_run:
        print("  [Status_Components] --dry-run: файл не записан.")
        return

    df["lease_restricted"] = new_lease
    df.to_excel(path, engine="openpyxl", index=False)
    print(f"  [Status_Components] Записано: {path}")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Нормализация папки source_data перед extract (Program_AC, Status_Components)."
    )
    parser.add_argument(
        "--dataset",
        required=True,
        help="Путь к папке v_YYYY-MM-DD (абсолютный или от корня репозитория)",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Только вывести план действий, без записи файлов",
    )
    parser.add_argument(
        "--sync-lease",
        action="store_true",
        help="Если колонка lease_restricted уже есть — пересчитать её целиком по owner",
    )
    args = parser.parse_args()

    root = resolve_dataset_path(args.dataset)
    if not root.is_dir():
        _fail(f"Не каталог или не найдено: {root}")

    program_ac = root / PROGRAM_AC_NAME
    status_sc = root / STATUS_COMPONENTS_NAME

    if not program_ac.is_file():
        _fail(f"Нет обязательного файла: {program_ac}")
    if not status_sc.is_file():
        _fail(f"Нет обязательного файла: {status_sc}")

    print(f"Датасет: {root}")
    if args.dry_run:
        print("Режим: --dry-run (без записи)")
    normalize_program_ac_header(program_ac, dry_run=args.dry_run)
    normalize_status_lease_restricted(
        status_sc, dry_run=args.dry_run, sync_lease=args.sync_lease
    )
    print("Готово (exit 0).")


if __name__ == "__main__":
    main()
